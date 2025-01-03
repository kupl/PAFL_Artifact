import os
import sys
import json
from openpyxl import Workbook

cpp_project = [
    "cpp_peglib",
    "cppcheck",
    "exiv2",
    "libchewing",
    "libxml2",
    "proj",
    "openssl",
    "yaml_cpp",
]
python_project = ["thefuck", "fastapi", "spacy", "youtube-dl"]


def interval(beg: int, end: int) -> list:
    return list(range(beg, end + 1)) if beg < end else list(range(beg, end - 1, -1))


version_map: dict[str, int] = dict()
version_map["cppcheck"] = [
    7,
    1,
    27,
    2,
    15,
    29,
    14,
    6,
    23,
    5,
    17,
    30,
    24,
    8,
    16,
    25,
    3,
    19,
    9,
    18,
    12,
    11,
    21,
    28,
    13,
    20,
    22,
    4,
    26,
    10,
]
version_map["proj"] = [
    24,
    8,
    13,
    12,
    11,
    21,
    22,
    9,
    18,
    16,
    1,
    3,
    17,
    2,
    4,
    19,
    10,
    6,
    20,
    5,
    27,
    23,
    14,
    28,
    15,
    7,
    26,
    25,
]
version_map["openssl"] = [
    13,
    1,
    2,
    9,
    12,
    19,
    14,
    28,
    11,
    15,
    5,
    18,
    16,
    22,
    10,
    23,
    26,
    24,
    25,
    17,
    6,
    4,
    7,
    27,
    20,
    21,
]
version_map["cpp_peglib"] = [8, 9, 7, 6, 5, 4, 2, 3, 1, 10]
version_map["exiv2"] = [
    10,
    15,
    14,
    4,
    13,
    6,
    8,
    9,
    11,
    12,
    7,
    5,
    16,
    17,
    3,
    2,
    1,
    18,
    19,
    20,
]
version_map["libchewing"] = [6, 8, 3, 4, 5, 7, 2, 1]
version_map["libxml2"] = [4, 7, 1, 2, 3, 5, 6]
version_map["yaml_cpp"] = [10] + list(range(1, 9 + 1))
version_map["thefuck"] = interval(32, 13) + interval(11, 1)
version_map["fastapi"] = interval(16, 2)
version_map["spacy"] = interval(7, 1)
version_map["youtube-dl"] = (
    interval(35, 31)
    + interval(10, 8)
    + [30, 7, 29, 28, 6, 27, 26, 5, 25, 24]
    + interval(23, 18)
    + [3, 16, 2, 15, 1, 13, 12, 11]
)

proj = sys.argv[1]
method = sys.argv[2]
coverage_dir = f"/workspace/data/coverage/{proj}"
oracle_dir = coverage_dir + "/oracle"

# Init workbook(base)
workbook_base = Workbook()
sheet_base = workbook_base.active
sheet_base.cell(1, 1, "version")
sheet_base.cell(1, 2, "FR")
sheet_base.cell(1, 3, "AR")
sheet_base.cell(1, 4, "LOC")
sheet_base.cell(1, 5, "time(sec)")

# Init workbook(pafl)
workbook_pafl = Workbook()
sheet_pafl = workbook_pafl.active
sheet_pafl.cell(1, 1, "version")
sheet_pafl.cell(1, 2, "FR")
sheet_pafl.cell(1, 3, "AR")
sheet_pafl.cell(1, 4, "LOC")

# Print profile
print(f"{proj}-{method}")

# PAFL
for idx, ver in enumerate(version_map[proj], 1):
    base_dir = f"{coverage_dir}/buggy-{ver}/__pafl__/{proj}-{method}-base"
    pafl_dir = f"{coverage_dir}/buggy-{ver}/__pafl__/{proj}-{method}-pafl"

    # version
    sheet_base.cell(1 + idx, 1, ver)
    sheet_pafl.cell(1 + idx, 1, ver)

    # Read faults
    faults: list[tuple[str, int]] = list()
    with open(f"{base_dir}/ranking.json", "r") as base_file:
        base_json = json.load(base_file)
    base_ranking = list()
    for line in base_json["lines"]:
        if float(line["sus"]) > 0.0:
            base_ranking.append((line["file"], line["line"]))
    with open(f"{oracle_dir}/{ver}.json", "r") as oracle_file:
        oracle = json.load(oracle_file)
    for obj in oracle["locations"]:
        for line in obj["lines"]:
            item = (obj["file"], line)
            if item in base_ranking:
                faults.append(item)

    # Read ranking
    with open(f"{base_dir}/ranking.json", "r") as base_ranking:
        base_json = json.load(base_ranking)
    with open(f"{pafl_dir}/ranking.json", "r") as pafl_ranking:
        pafl_json = json.load(pafl_ranking)

    # total
    sheet_base.cell(1 + idx, 4, base_json["total"])
    sheet_pafl.cell(1 + idx, 4, pafl_json["total"])

    # FR, AR of base
    fr, ar, ar_size = -1, 0.0, 0
    for line in base_json["lines"]:
        if (line["file"], line["line"]) in faults:
            if fr == -1:
                fr = int(line["ranking"])
            ar += int(line["ranking"])
            ar_size += 1
    sheet_base.cell(1 + idx, 2, fr if fr != -1 else base_json["total"])
    sheet_base.cell(1 + idx, 3, ar / ar_size if ar_size != 0 else base_json["total"])

    # FR, AR of pafl
    fr, ar, ar_size = -1, 0.0, 0
    for line in pafl_json["lines"]:
        if (line["file"], line["line"]) in faults:
            if fr == -1:
                fr = int(line["ranking"])
            ar += int(line["ranking"])
            ar_size += 1
    sheet_pafl.cell(1 + idx, 2, fr if fr != -1 else pafl_json["total"])
    sheet_pafl.cell(1 + idx, 3, ar / ar_size if ar_size != 0 else pafl_json["total"])

# Save workbook
save_dir = "/workspace/evaluation"
if not os.path.exists(save_dir):
    os.makedirs(save_dir)
workbook_base.save(f"{save_dir}/{proj}-{method}-base.xlsx")
workbook_pafl.save(f"{save_dir}/{proj}-{method}-pafl.xlsx")
